#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
仓库管理系统 - Excel报表版本
功能：管理仓库货物出入库、库存跟踪、供应商管理，并生成Excel报表
作者：AI Assistant
日期：2024
"""

import sqlite3
import datetime
import pandas as pd
import os
from typing import List, Dict, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class WarehouseManagementSystemExcel:
    """仓库管理系统主类 - Excel报表版本"""
    
    def __init__(self, db_path: str = "warehouse.db", excel_path: str = "warehouse_report.xlsx"):
        """
        初始化仓库管理系统
        
        Args:
            db_path: 数据库文件路径
            excel_path: Excel报表文件路径
        """
        self.db_path = db_path
        self.excel_path = excel_path
        self.conn = None
        self.cursor = None
        
    def connect_database(self):
        """连接数据库"""
        try:
            self.conn = sqlite3.connect(self.db_path)
            self.cursor = self.conn.cursor()
            print("✅ 数据库连接成功")
            return True
        except Exception as e:
            print(f"❌ 数据库连接失败: {e}")
            return False
    
    def close_database(self):
        """关闭数据库连接"""
        if self.conn:
            self.conn.close()
            print("🔒 数据库连接已关闭")
    
    def create_tables(self):
        """创建数据库表结构（简化版，适配SQLite）"""
        try:
            # 操作员表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS caozuoyuan (
                    xingming VARCHAR(20) PRIMARY KEY,
                    caozuoyuanlianxifangshi VARCHAR(20)
                )
            ''')
            
            # 供应商表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS gongyingshang (
                    gongyingshangbianhao VARCHAR(20) PRIMARY KEY,
                    gongyingshangmingcheng VARCHAR(20),
                    lianxirren VARCHAR(20),
                    lianxifangshi VARCHAR(20)
                )
            ''')
            
            # 仓库表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS cangku (
                    cangkumingcheng VARCHAR(20) PRIMARY KEY,
                    xingming VARCHAR(20),
                    cangkufuzeren VARCHAR(20),
                    cangkuchuangjianriqi VARCHAR(20),
                    FOREIGN KEY (xingming) REFERENCES caozuoyuan (xingming)
                )
            ''')
            
            # 库存表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS kucun (
                    bianhao VARCHAR(20) PRIMARY KEY,
                    cangkumingcheng VARCHAR(20),
                    shuliang INTEGER DEFAULT 0,
                    danjia DECIMAL(10,2) DEFAULT 0.00,
                    FOREIGN KEY (cangkumingcheng) REFERENCES cangku (cangkumingcheng)
                )
            ''')
            
            # 入库表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS ruku (
                    rukubianhao VARCHAR(20) PRIMARY KEY,
                    bianhao VARCHAR(20),
                    huowubianhao VARCHAR(20),
                    shuliang INTEGER,
                    mingcheng VARCHAR(20),
                    rukuriqi VARCHAR(20),
                    danjia DECIMAL(10,2),
                    gongyingshangmingcheng VARCHAR(20),
                    FOREIGN KEY (bianhao) REFERENCES kucun (bianhao)
                )
            ''')
            
            # 出库表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS chuku (
                    chukubianhao VARCHAR(20) PRIMARY KEY,
                    bianhao VARCHAR(20),
                    huowubianhao VARCHAR(20),
                    shuliang INTEGER,
                    mingcheng VARCHAR(20),
                    chukuriqi VARCHAR(20),
                    danjia DECIMAL(10,2),
                    FOREIGN KEY (bianhao) REFERENCES kucun (bianhao)
                )
            ''')
            
            # 供应关系表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS gongying (
                    gongyingshangbianhao VARCHAR(20),
                    cangkumingcheng VARCHAR(20),
                    PRIMARY KEY (gongyingshangbianhao, cangkumingcheng),
                    FOREIGN KEY (gongyingshangbianhao) REFERENCES gongyingshang (gongyingshangbianhao),
                    FOREIGN KEY (cangkumingcheng) REFERENCES cangku (cangkumingcheng)
                )
            ''')
            
            self.conn.commit()
            print("✅ 数据库表创建成功")
            return True
            
        except Exception as e:
            print(f"❌ 创建表失败: {e}")
            return False
    
    def insert_sample_data(self):
        """插入示例数据"""
        try:
            # 插入操作员数据
            operators = [
                ("张三", "13800138001"),
                ("李四", "13800138002"),
                ("王五", "13800138003")
            ]
            self.cursor.executemany(
                "INSERT OR REPLACE INTO caozuoyuan VALUES (?, ?)", 
                operators
            )
            
            # 插入供应商数据
            suppliers = [
                ("SP001", "北京电子有限公司", "张经理", "010-12345678"),
                ("SP002", "上海机械制造厂", "李经理", "021-87654321"),
                ("SP003", "广州贸易公司", "王经理", "020-11223344")
            ]
            self.cursor.executemany(
                "INSERT OR REPLACE INTO gongyingshang VALUES (?, ?, ?, ?)", 
                suppliers
            )
            
            # 插入仓库数据
            warehouses = [
                ("主仓库", "张三", "张主任", "2023-01-01"),
                ("分仓库A", "李四", "李主任", "2023-02-01"),
                ("分仓库B", "王五", "王主任", "2023-03-01")
            ]
            self.cursor.executemany(
                "INSERT OR REPLACE INTO cangku VALUES (?, ?, ?, ?)", 
                warehouses
            )
            
            # 插入库存数据
            inventory = [
                ("INV001", "主仓库", 100, 50.00),
                ("INV002", "主仓库", 200, 30.00),
                ("INV003", "分仓库A", 150, 80.00),
                ("INV004", "分仓库B", 80, 120.00)
            ]
            self.cursor.executemany(
                "INSERT OR REPLACE INTO kucun VALUES (?, ?, ?, ?)", 
                inventory
            )
            
            # 插入供应关系数据
            supply_relations = [
                ("SP001", "主仓库"),
                ("SP002", "主仓库"),
                ("SP002", "分仓库A"),
                ("SP003", "分仓库B")
            ]
            self.cursor.executemany(
                "INSERT OR REPLACE INTO gongying VALUES (?, ?)", 
                supply_relations
            )
            
            self.conn.commit()
            print("✅ 示例数据插入成功")
            return True
            
        except Exception as e:
            print(f"❌ 插入示例数据失败: {e}")
            return False
    
    def get_all_data_for_excel(self) -> Dict[str, pd.DataFrame]:
        """获取所有数据用于Excel报表"""
        data = {}
        
        try:
            # 操作员数据
            self.cursor.execute('SELECT * FROM caozuoyuan')
            operators_data = self.cursor.fetchall()
            data['操作员'] = pd.DataFrame(operators_data, columns=['姓名', '联系方式'])
            
            # 供应商数据
            self.cursor.execute('SELECT * FROM gongyingshang')
            suppliers_data = self.cursor.fetchall()
            data['供应商'] = pd.DataFrame(suppliers_data, columns=['供应商编号', '供应商名称', '联系人', '联系方式'])
            
            # 仓库数据
            self.cursor.execute('SELECT * FROM cangku')
            warehouses_data = self.cursor.fetchall()
            data['仓库'] = pd.DataFrame(warehouses_data, columns=['仓库名称', '操作员', '负责人', '创建日期'])
            
            # 库存数据
            self.cursor.execute('''
                SELECT k.bianhao, k.cangkumingcheng, k.shuliang, k.danjia,
                       c.cangkufuzeren, (k.shuliang * k.danjia) as 总价值
                FROM kucun k
                LEFT JOIN cangku c ON k.cangkumingcheng = c.cangkumingcheng
                ORDER BY k.cangkumingcheng, k.bianhao
            ''')
            inventory_data = self.cursor.fetchall()
            data['库存'] = pd.DataFrame(inventory_data, columns=['库存编号', '仓库名称', '数量', '单价', '负责人', '总价值'])
            
            # 入库记录
            self.cursor.execute('''
                SELECT r.rukubianhao, r.huowubianhao, r.mingcheng, r.shuliang, 
                       r.danjia, r.rukuriqi, r.gongyingshangmingcheng,
                       (r.shuliang * r.danjia) as 入库金额
                FROM ruku r
                ORDER BY r.rukuriqi DESC
            ''')
            inbound_data = self.cursor.fetchall()
            data['入库记录'] = pd.DataFrame(inbound_data, columns=['入库编号', '货物编号', '货物名称', '数量', '单价', '入库日期', '供应商', '入库金额'])
            
            # 出库记录
            self.cursor.execute('''
                SELECT c.chukubianhao, c.huowubianhao, c.mingcheng, c.shuliang, 
                       c.danjia, c.chukuriqi, (c.shuliang * c.danjia) as 出库金额
                FROM chuku c
                ORDER BY c.chukuriqi DESC
            ''')
            outbound_data = self.cursor.fetchall()
            data['出库记录'] = pd.DataFrame(outbound_data, columns=['出库编号', '货物编号', '货物名称', '数量', '单价', '出库日期', '出库金额'])
            
            # 仓库汇总
            self.cursor.execute('''
                SELECT c.cangkumingcheng, c.cangkufuzeren, c.xingming,
                       COUNT(k.bianhao) as 库存种类,
                       SUM(k.shuliang) as 总数量,
                       SUM(k.shuliang * k.danjia) as 总价值
                FROM cangku c
                LEFT JOIN kucun k ON c.cangkumingcheng = k.cangkumingcheng
                GROUP BY c.cangkumingcheng, c.cangkufuzeren, c.xingming
                ORDER BY c.cangkumingcheng
            ''')
            summary_data = self.cursor.fetchall()
            data['仓库汇总'] = pd.DataFrame(summary_data, columns=['仓库名称', '负责人', '操作员', '库存种类', '总数量', '总价值'])
            
            # 供应关系
            self.cursor.execute('''
                SELECT g.gongyingshangbianhao, s.gongyingshangmingcheng, 
                       g.cangkumingcheng, s.lianxirren, s.lianxifangshi
                FROM gongying g
                LEFT JOIN gongyingshang s ON g.gongyingshangbianhao = s.gongyingshangbianhao
                ORDER BY g.gongyingshangbianhao, g.cangkumingcheng
            ''')
            supply_data = self.cursor.fetchall()
            data['供应关系'] = pd.DataFrame(supply_data, columns=['供应商编号', '供应商名称', '仓库名称', '联系人', '联系方式'])
            
            return data
            
        except Exception as e:
            print(f"❌ 获取数据失败: {e}")
            return {}
    
    def generate_excel_report(self, operation_name: str = ""):
        """生成Excel报表"""
        try:
            # 获取所有数据
            data = self.get_all_data_for_excel()
            
            # 创建Excel工作簿
            wb = Workbook()
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 创建报表信息工作表
            info_sheet = wb.active
            info_sheet.title = "报表信息"
            
            # 添加报表信息
            info_sheet['A1'] = "仓库管理系统 - Excel报表"
            info_sheet['A1'].font = Font(bold=True, size=16)
            info_sheet['A3'] = f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            info_sheet['A4'] = f"操作类型: {operation_name}" if operation_name else "操作类型: 系统状态查看"
            info_sheet['A5'] = f"数据库文件: {self.db_path}"
            
            # 为每个数据表创建工作表
            for sheet_name, df in data.items():
                if not df.empty:
                    # 创建工作表
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # 写入数据
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    
                    # 设置表头样式
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # 设置列宽
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # 添加边框
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for row in ws.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                        for cell in row:
                            cell.border = thin_border
            
            # 保存Excel文件
            wb.save(self.excel_path)
            print(f"✅ Excel报表已生成: {self.excel_path}")
            
            # 显示文件信息
            file_size = os.path.getsize(self.excel_path) / 1024  # KB
            print(f"📊 文件大小: {file_size:.2f} KB")
            
            return True
            
        except Exception as e:
            print(f"❌ 生成Excel报表失败: {e}")
            return False
    
    def add_operator(self, name: str, contact: str) -> bool:
        """添加操作员"""
        try:
            self.cursor.execute(
                "INSERT INTO caozuoyuan VALUES (?, ?)", 
                (name, contact)
            )
            self.conn.commit()
            print(f"✅ 操作员 {name} 添加成功")
            self.generate_excel_report(f"添加操作员: {name}")
            return True
        except Exception as e:
            print(f"❌ 添加操作员失败: {e}")
            return False
    
    def add_warehouse(self, name: str, operator: str, manager: str) -> bool:
        """添加仓库"""
        try:
            create_date = datetime.datetime.now().strftime("%Y-%m-%d")
            self.cursor.execute(
                "INSERT INTO cangku VALUES (?, ?, ?, ?)", 
                (name, operator, manager, create_date)
            )
            self.conn.commit()
            print(f"✅ 仓库 {name} 添加成功")
            self.generate_excel_report(f"添加仓库: {name}")
            return True
        except Exception as e:
            print(f"❌ 添加仓库失败: {e}")
            return False
    
    def add_inventory(self, code: str, warehouse: str, quantity: int, price: float) -> bool:
        """添加库存"""
        try:
            self.cursor.execute(
                "INSERT INTO kucun VALUES (?, ?, ?, ?)", 
                (code, warehouse, quantity, price)
            )
            self.conn.commit()
            print(f"✅ 库存 {code} 添加成功")
            self.generate_excel_report(f"添加库存: {code}")
            return True
        except Exception as e:
            print(f"❌ 添加库存失败: {e}")
            return False
    
    def process_inbound(self, inbound_code: str, inventory_code: str, 
                       goods_code: str, quantity: int, name: str, 
                       price: float, supplier: str) -> bool:
        """处理入库操作"""
        try:
            # 记录入库信息
            inbound_date = datetime.datetime.now().strftime("%Y-%m-%d")
            self.cursor.execute(
                "INSERT INTO ruku VALUES (?, ?, ?, ?, ?, ?, ?, ?)", 
                (inbound_code, inventory_code, goods_code, quantity, 
                 name, inbound_date, price, supplier)
            )
            
            # 更新库存数量
            self.cursor.execute(
                "UPDATE kucun SET shuliang = shuliang + ? WHERE bianhao = ?", 
                (quantity, inventory_code)
            )
            
            self.conn.commit()
            print(f"✅ 入库操作 {inbound_code} 处理成功")
            self.generate_excel_report(f"入库操作: {inbound_code}")
            return True
        except Exception as e:
            print(f"❌ 入库操作失败: {e}")
            return False
    
    def process_outbound(self, outbound_code: str, inventory_code: str,
                        goods_code: str, quantity: int, name: str, price: float) -> bool:
        """处理出库操作"""
        try:
            # 检查库存是否足够
            self.cursor.execute(
                "SELECT shuliang FROM kucun WHERE bianhao = ?", 
                (inventory_code,)
            )
            current_quantity = self.cursor.fetchone()
            
            if not current_quantity or current_quantity[0] < quantity:
                print(f"❌ 库存不足，当前库存: {current_quantity[0] if current_quantity else 0}, 需要: {quantity}")
                return False
            
            # 记录出库信息
            outbound_date = datetime.datetime.now().strftime("%Y-%m-%d")
            self.cursor.execute(
                "INSERT INTO chuku VALUES (?, ?, ?, ?, ?, ?, ?)", 
                (outbound_code, inventory_code, goods_code, quantity, 
                 name, outbound_date, price)
            )
            
            # 更新库存数量
            self.cursor.execute(
                "UPDATE kucun SET shuliang = shuliang - ? WHERE bianhao = ?", 
                (quantity, inventory_code)
            )
            
            self.conn.commit()
            print(f"✅ 出库操作 {outbound_code} 处理成功")
            self.generate_excel_report(f"出库操作: {outbound_code}")
            return True
        except Exception as e:
            print(f"❌ 出库操作失败: {e}")
            return False
    
    def print_inventory_status(self):
        """打印库存状态"""
        print("\n" + "="*60)
        print("📊 当前库存状态")
        print("="*60)
        
        try:
            self.cursor.execute('''
                SELECT k.bianhao, k.cangkumingcheng, k.shuliang, k.danjia,
                       c.cangkufuzeren, (k.shuliang * k.danjia) as 总价值
                FROM kucun k
                LEFT JOIN cangku c ON k.cangkumingcheng = c.cangkumingcheng
                ORDER BY k.cangkumingcheng, k.bianhao
            ''')
            
            results = self.cursor.fetchall()
            if results:
                print(f"{'库存编号':<12} {'仓库名称':<12} {'数量':<8} {'单价':<10} {'负责人':<10} {'总价值':<12}")
                print("-" * 70)
                for row in results:
                    print(f"{row[0]:<12} {row[1]:<12} {row[2]:<8} {row[3]:<10} {row[4]:<10} {row[5]:<12}")
            else:
                print("暂无库存数据")
        except Exception as e:
            print(f"❌ 获取库存状态失败: {e}")
    
    def print_warehouse_summary(self):
        """打印仓库汇总信息"""
        print("\n" + "="*80)
        print("🏢 仓库汇总信息")
        print("="*80)
        
        try:
            self.cursor.execute('''
                SELECT c.cangkumingcheng, c.cangkufuzeren, c.xingming,
                       COUNT(k.bianhao) as 库存种类,
                       SUM(k.shuliang) as 总数量,
                       SUM(k.shuliang * k.danjia) as 总价值
                FROM cangku c
                LEFT JOIN kucun k ON c.cangkumingcheng = k.cangkumingcheng
                GROUP BY c.cangkumingcheng, c.cangkufuzeren, c.xingming
                ORDER BY c.cangkumingcheng
            ''')
            
            results = self.cursor.fetchall()
            if results:
                print(f"{'仓库名称':<12} {'负责人':<10} {'操作员':<10} {'库存种类':<8} {'总数量':<8} {'总价值':<12}")
                print("-" * 80)
                for row in results:
                    print(f"{row[0]:<12} {row[1]:<10} {row[2]:<10} {row[3]:<8} {row[4]:<8} {row[5]:<12}")
            else:
                print("暂无仓库数据")
        except Exception as e:
            print(f"❌ 获取仓库汇总失败: {e}")

def main():
    """主函数 - 演示仓库管理系统的使用"""
    print("🚀 仓库管理系统 - Excel报表版本")
    print("="*60)
    
    # 创建系统实例
    wms = WarehouseManagementSystemExcel()
    
    # 连接数据库
    if not wms.connect_database():
        return
    
    # 创建表结构
    if not wms.create_tables():
        wms.close_database()
        return
    
    # 插入示例数据
    if not wms.insert_sample_data():
        wms.close_database()
        return
    
    # 生成初始Excel报表
    print("\n📊 生成初始Excel报表...")
    wms.generate_excel_report("系统初始化")
    
    # 显示初始状态
    wms.print_inventory_status()
    wms.print_warehouse_summary()
    
    # 演示入库操作
    print("\n" + "="*60)
    print("📥 演示入库操作")
    print("="*60)
    
    wms.process_inbound(
        inbound_code="IN001",
        inventory_code="INV001",
        goods_code="GOODS001",
        quantity=50,
        name="笔记本电脑",
        price=5000.00,
        supplier="北京电子有限公司"
    )
    
    wms.process_inbound(
        inbound_code="IN002",
        inventory_code="INV002",
        goods_code="GOODS002",
        quantity=100,
        name="办公椅",
        price=300.00,
        supplier="上海机械制造厂"
    )
    
    # 演示出库操作
    print("\n" + "="*60)
    print("📤 演示出库操作")
    print("="*60)
    
    wms.process_outbound(
        outbound_code="OUT001",
        inventory_code="INV001",
        goods_code="GOODS001",
        quantity=20,
        name="笔记本电脑",
        price=5000.00
    )
    
    wms.process_outbound(
        outbound_code="OUT002",
        inventory_code="INV002",
        goods_code="GOODS002",
        quantity=30,
        name="办公椅",
        price=300.00
    )
    
    # 显示操作后的状态
    print("\n" + "="*60)
    print("📊 操作后的状态")
    print("="*60)
    
    wms.print_inventory_status()
    wms.print_warehouse_summary()
    
    # 生成最终Excel报表
    print("\n📊 生成最终Excel报表...")
    wms.generate_excel_report("操作完成")
    
    # 关闭数据库连接
    wms.close_database()
    
    print("\n" + "="*60)
    print("✅ 仓库管理系统演示完成")
    print(f"📄 Excel报表文件: {wms.excel_path}")
    print("="*60)

if __name__ == "__main__":
    main() 