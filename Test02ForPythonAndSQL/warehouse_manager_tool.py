#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
仓库管理系统 - 便捷操作工具
功能：创建空白数据库、Excel文档，提供便捷的入库出库操作界面
作者：AI Assistant
日期：2024
"""

import sqlite3
import datetime
import pandas as pd
import os
import sys
from typing import List, Dict, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class WarehouseManagerTool:
    """仓库管理便捷工具"""
    
    def __init__(self, db_path: str = "warehouse.db", excel_path: str = "warehouse_report.xlsx"):
        """
        初始化仓库管理工具
        
        Args:
            db_path: 数据库文件路径
            excel_path: Excel报表文件路径
        """
        self.db_path = db_path
        self.excel_path = excel_path
        self.conn = None
        self.cursor = None
        
    def create_blank_database(self):
        """创建空白数据库"""
        try:
            # 如果数据库文件已存在，先删除
            if os.path.exists(self.db_path):
                os.remove(self.db_path)
                print(f"🗑️ 删除旧数据库文件: {self.db_path}")
            
            # 连接数据库（会自动创建文件）
            self.conn = sqlite3.connect(self.db_path)
            self.cursor = self.conn.cursor()
            
            # 创建表结构
            self.create_tables()
            
            print(f"✅ 空白数据库创建成功: {self.db_path}")
            return True
            
        except Exception as e:
            print(f"❌ 创建空白数据库失败: {e}")
            return False
    
    def create_blank_excel(self):
        """创建空白Excel文档"""
        try:
            # 如果Excel文件已存在，先删除
            if os.path.exists(self.excel_path):
                os.remove(self.excel_path)
                print(f"🗑️ 删除旧Excel文件: {self.excel_path}")
            
            # 创建Excel工作簿
            wb = Workbook()
            
            # 创建各个工作表
            sheets = [
                "操作员", "供应商", "仓库", "库存", 
                "入库记录", "出库记录", "仓库汇总", "供应关系"
            ]
            
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 创建新的工作表
            for sheet_name in sheets:
                ws = wb.create_sheet(title=sheet_name)
                
                # 设置表头样式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # 根据工作表类型设置表头
                if sheet_name == "操作员":
                    headers = ["姓名", "联系方式"]
                elif sheet_name == "供应商":
                    headers = ["供应商编号", "供应商名称", "联系人", "联系方式"]
                elif sheet_name == "仓库":
                    headers = ["仓库名称", "操作员", "负责人", "创建日期"]
                elif sheet_name == "库存":
                    headers = ["库存编号", "仓库名称", "数量", "单价", "负责人", "总价值"]
                elif sheet_name == "入库记录":
                    headers = ["入库编号", "货物编号", "货物名称", "数量", "单价", "入库日期", "供应商", "入库金额"]
                elif sheet_name == "出库记录":
                    headers = ["出库编号", "货物编号", "货物名称", "数量", "单价", "出库日期", "出库金额"]
                elif sheet_name == "仓库汇总":
                    headers = ["仓库名称", "负责人", "操作员", "库存种类", "总数量", "总价值"]
                elif sheet_name == "供应关系":
                    headers = ["供应商编号", "供应商名称", "仓库名称", "联系人", "联系方式"]
                
                # 写入表头
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 设置列宽
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[chr(64 + col)].width = 15
            
            # 保存Excel文件
            wb.save(self.excel_path)
            print(f"✅ 空白Excel文档创建成功: {self.excel_path}")
            return True
            
        except Exception as e:
            print(f"❌ 创建空白Excel文档失败: {e}")
            return False
    
    def create_tables(self):
        """创建数据库表结构"""
        try:
            # 操作员表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS caozuoyuan (
                    xingming VARCHAR(20) PRIMARY KEY,
                    caozuoyuanlianxifangshi VARCHAR(20)
                )
            ''')
            
            # 供应商表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS gongyingshang (
                    gongyingshangbianhao VARCHAR(20) PRIMARY KEY,
                    gongyingshangmingcheng VARCHAR(20),
                    lianxirren VARCHAR(20),
                    lianxifangshi VARCHAR(20)
                )
            ''')
            
            # 仓库表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS cangku (
                    cangkumingcheng VARCHAR(20) PRIMARY KEY,
                    xingming VARCHAR(20),
                    cangkufuzeren VARCHAR(20),
                    cangkuchuangjianriqi VARCHAR(20),
                    FOREIGN KEY (xingming) REFERENCES caozuoyuan (xingming)
                )
            ''')
            
            # 库存表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS kucun (
                    bianhao VARCHAR(20) PRIMARY KEY,
                    cangkumingcheng VARCHAR(20),
                    shuliang INTEGER DEFAULT 0,
                    danjia DECIMAL(10,2) DEFAULT 0.00,
                    FOREIGN KEY (cangkumingcheng) REFERENCES cangku (cangkumingcheng)
                )
            ''')
            
            # 入库表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS ruku (
                    rukubianhao VARCHAR(20) PRIMARY KEY,
                    bianhao VARCHAR(20),
                    huowubianhao VARCHAR(20),
                    shuliang INTEGER,
                    mingcheng VARCHAR(20),
                    rukuriqi VARCHAR(20),
                    danjia DECIMAL(10,2),
                    gongyingshangmingcheng VARCHAR(20),
                    FOREIGN KEY (bianhao) REFERENCES kucun (bianhao)
                )
            ''')
            
            # 出库表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS chuku (
                    chukubianhao VARCHAR(20) PRIMARY KEY,
                    bianhao VARCHAR(20),
                    huowubianhao VARCHAR(20),
                    shuliang INTEGER,
                    mingcheng VARCHAR(20),
                    chukuriqi VARCHAR(20),
                    danjia DECIMAL(10,2),
                    FOREIGN KEY (bianhao) REFERENCES kucun (bianhao)
                )
            ''')
            
            # 供应关系表
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS gongying (
                    gongyingshangbianhao VARCHAR(20),
                    cangkumingcheng VARCHAR(20),
                    PRIMARY KEY (gongyingshangbianhao, cangkumingcheng),
                    FOREIGN KEY (gongyingshangbianhao) REFERENCES gongyingshang (gongyingshangbianhao),
                    FOREIGN KEY (cangkumingcheng) REFERENCES cangku (cangkumingcheng)
                )
            ''')
            
            self.conn.commit()
            print("✅ 数据库表结构创建成功")
            return True
            
        except Exception as e:
            print(f"❌ 创建表结构失败: {e}")
            return False
    
    def connect_database(self):
        """连接数据库"""
        try:
            self.conn = sqlite3.connect(self.db_path)
            self.cursor = self.conn.cursor()
            print("✅ 数据库连接成功")
            return True
        except Exception as e:
            print(f"❌ 数据库连接失败: {e}")
            return False
    
    def close_database(self):
        """关闭数据库连接"""
        if self.conn:
            self.conn.close()
            print("🔒 数据库连接已关闭")
    
    def update_excel_report(self, operation_name: str = ""):
        """更新Excel报表"""
        try:
            # 获取所有数据
            data = self.get_all_data_for_excel()
            
            # 创建Excel工作簿
            wb = Workbook()
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 创建报表信息工作表
            info_sheet = wb.active
            info_sheet.title = "报表信息"
            
            # 添加报表信息
            info_sheet['A1'] = "仓库管理系统 - Excel报表"
            info_sheet['A1'].font = Font(bold=True, size=16)
            info_sheet['A3'] = f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            info_sheet['A4'] = f"操作类型: {operation_name}" if operation_name else "操作类型: 系统状态查看"
            info_sheet['A5'] = f"数据库文件: {self.db_path}"
            
            # 为每个数据表创建工作表
            for sheet_name, df in data.items():
                if not df.empty:
                    # 创建工作表
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # 写入数据
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    
                    # 设置表头样式
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # 设置列宽
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # 添加边框
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for row in ws.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                        for cell in row:
                            cell.border = thin_border
            
            # 保存Excel文件
            wb.save(self.excel_path)
            print(f"✅ Excel报表已更新: {self.excel_path}")
            
            return True
            
        except Exception as e:
            print(f"❌ 更新Excel报表失败: {e}")
            return False
    
    def get_all_data_for_excel(self) -> Dict[str, pd.DataFrame]:
        """获取所有数据用于Excel报表"""
        data = {}
        
        try:
            # 操作员数据
            self.cursor.execute('SELECT * FROM caozuoyuan')
            operators_data = self.cursor.fetchall()
            data['操作员'] = pd.DataFrame(operators_data, columns=['姓名', '联系方式'])
            
            # 供应商数据
            self.cursor.execute('SELECT * FROM gongyingshang')
            suppliers_data = self.cursor.fetchall()
            data['供应商'] = pd.DataFrame(suppliers_data, columns=['供应商编号', '供应商名称', '联系人', '联系方式'])
            
            # 仓库数据
            self.cursor.execute('SELECT * FROM cangku')
            warehouses_data = self.cursor.fetchall()
            data['仓库'] = pd.DataFrame(warehouses_data, columns=['仓库名称', '操作员', '负责人', '创建日期'])
            
            # 库存数据
            self.cursor.execute('''
                SELECT k.bianhao, k.cangkumingcheng, k.shuliang, k.danjia,
                       c.cangkufuzeren, (k.shuliang * k.danjia) as 总价值
                FROM kucun k
                LEFT JOIN cangku c ON k.cangkumingcheng = c.cangkumingcheng
                ORDER BY k.cangkumingcheng, k.bianhao
            ''')
            inventory_data = self.cursor.fetchall()
            data['库存'] = pd.DataFrame(inventory_data, columns=['库存编号', '仓库名称', '数量', '单价', '负责人', '总价值'])
            
            # 入库记录
            self.cursor.execute('''
                SELECT r.rukubianhao, r.huowubianhao, r.mingcheng, r.shuliang, 
                       r.danjia, r.rukuriqi, r.gongyingshangmingcheng,
                       (r.shuliang * r.danjia) as 入库金额
                FROM ruku r
                ORDER BY r.rukuriqi DESC
            ''')
            inbound_data = self.cursor.fetchall()
            data['入库记录'] = pd.DataFrame(inbound_data, columns=['入库编号', '货物编号', '货物名称', '数量', '单价', '入库日期', '供应商', '入库金额'])
            
            # 出库记录
            self.cursor.execute('''
                SELECT c.chukubianhao, c.huowubianhao, c.mingcheng, c.shuliang, 
                       c.danjia, c.chukuriqi, (c.shuliang * c.danjia) as 出库金额
                FROM chuku c
                ORDER BY c.chukuriqi DESC
            ''')
            outbound_data = self.cursor.fetchall()
            data['出库记录'] = pd.DataFrame(outbound_data, columns=['出库编号', '货物编号', '货物名称', '数量', '单价', '出库日期', '出库金额'])
            
            # 仓库汇总
            self.cursor.execute('''
                SELECT c.cangkumingcheng, c.cangkufuzeren, c.xingming,
                       COUNT(k.bianhao) as 库存种类,
                       SUM(k.shuliang) as 总数量,
                       SUM(k.shuliang * k.danjia) as 总价值
                FROM cangku c
                LEFT JOIN kucun k ON c.cangkumingcheng = k.cangkumingcheng
                GROUP BY c.cangkumingcheng, c.cangkufuzeren, c.xingming
                ORDER BY c.cangkumingcheng
            ''')
            summary_data = self.cursor.fetchall()
            data['仓库汇总'] = pd.DataFrame(summary_data, columns=['仓库名称', '负责人', '操作员', '库存种类', '总数量', '总价值'])
            
            # 供应关系
            self.cursor.execute('''
                SELECT g.gongyingshangbianhao, s.gongyingshangmingcheng, 
                       g.cangkumingcheng, s.lianxirren, s.lianxifangshi
                FROM gongying g
                LEFT JOIN gongyingshang s ON g.gongyingshangbianhao = s.gongyingshangbianhao
                ORDER BY g.gongyingshangbianhao, g.cangkumingcheng
            ''')
            supply_data = self.cursor.fetchall()
            data['供应关系'] = pd.DataFrame(supply_data, columns=['供应商编号', '供应商名称', '仓库名称', '联系人', '联系方式'])
            
            return data
            
        except Exception as e:
            print(f"❌ 获取数据失败: {e}")
            return {}
    
    def add_operator(self, name: str, contact: str) -> bool:
        """添加操作员"""
        try:
            self.cursor.execute(
                "INSERT INTO caozuoyuan VALUES (?, ?)", 
                (name, contact)
            )
            self.conn.commit()
            print(f"✅ 操作员 {name} 添加成功")
            self.update_excel_report(f"添加操作员: {name}")
            return True
        except Exception as e:
            print(f"❌ 添加操作员失败: {e}")
            return False
    
    def add_supplier(self, code: str, name: str, contact: str, phone: str) -> bool:
        """添加供应商"""
        try:
            self.cursor.execute(
                "INSERT INTO gongyingshang VALUES (?, ?, ?, ?)", 
                (code, name, contact, phone)
            )
            self.conn.commit()
            print(f"✅ 供应商 {name} 添加成功")
            self.update_excel_report(f"添加供应商: {name}")
            return True
        except Exception as e:
            print(f"❌ 添加供应商失败: {e}")
            return False
    
    def add_warehouse(self, name: str, operator: str, manager: str) -> bool:
        """添加仓库"""
        try:
            create_date = datetime.datetime.now().strftime("%Y-%m-%d")
            self.cursor.execute(
                "INSERT INTO cangku VALUES (?, ?, ?, ?)", 
                (name, operator, manager, create_date)
            )
            self.conn.commit()
            print(f"✅ 仓库 {name} 添加成功")
            self.update_excel_report(f"添加仓库: {name}")
            return True
        except Exception as e:
            print(f"❌ 添加仓库失败: {e}")
            return False
    
    def add_inventory(self, code: str, warehouse: str, quantity: int, price: float) -> bool:
        """添加库存"""
        try:
            self.cursor.execute(
                "INSERT INTO kucun VALUES (?, ?, ?, ?)", 
                (code, warehouse, quantity, price)
            )
            self.conn.commit()
            print(f"✅ 库存 {code} 添加成功")
            self.update_excel_report(f"添加库存: {code}")
            return True
        except Exception as e:
            print(f"❌ 添加库存失败: {e}")
            return False
    
    def process_inbound(self, inbound_code: str, inventory_code: str, 
                       goods_code: str, quantity: int, name: str, 
                       price: float, supplier: str) -> bool:
        """处理入库操作"""
        try:
            # 记录入库信息
            inbound_date = datetime.datetime.now().strftime("%Y-%m-%d")
            self.cursor.execute(
                "INSERT INTO ruku VALUES (?, ?, ?, ?, ?, ?, ?, ?)", 
                (inbound_code, inventory_code, goods_code, quantity, 
                 name, inbound_date, price, supplier)
            )
            
            # 更新库存数量
            self.cursor.execute(
                "UPDATE kucun SET shuliang = shuliang + ? WHERE bianhao = ?", 
                (quantity, inventory_code)
            )
            
            self.conn.commit()
            print(f"✅ 入库操作 {inbound_code} 处理成功")
            self.update_excel_report(f"入库操作: {inbound_code}")
            return True
        except Exception as e:
            print(f"❌ 入库操作失败: {e}")
            return False
    
    def process_outbound(self, outbound_code: str, inventory_code: str,
                        goods_code: str, quantity: int, name: str, price: float) -> bool:
        """处理出库操作"""
        try:
            # 检查库存是否足够
            self.cursor.execute(
                "SELECT shuliang FROM kucun WHERE bianhao = ?", 
                (inventory_code,)
            )
            current_quantity = self.cursor.fetchone()
            
            if not current_quantity or current_quantity[0] < quantity:
                print(f"❌ 库存不足，当前库存: {current_quantity[0] if current_quantity else 0}, 需要: {quantity}")
                return False
            
            # 记录出库信息
            outbound_date = datetime.datetime.now().strftime("%Y-%m-%d")
            self.cursor.execute(
                "INSERT INTO chuku VALUES (?, ?, ?, ?, ?, ?, ?)", 
                (outbound_code, inventory_code, goods_code, quantity, 
                 name, outbound_date, price)
            )
            
            # 更新库存数量
            self.cursor.execute(
                "UPDATE kucun SET shuliang = shuliang - ? WHERE bianhao = ?", 
                (quantity, inventory_code)
            )
            
            self.conn.commit()
            print(f"✅ 出库操作 {outbound_code} 处理成功")
            self.update_excel_report(f"出库操作: {outbound_code}")
            return True
        except Exception as e:
            print(f"❌ 出库操作失败: {e}")
            return False
    
    def show_menu(self):
        """显示操作菜单"""
        print("\n" + "="*60)
        print("🏢 仓库管理系统 - 便捷操作工具")
        print("="*60)
        print("1. 创建空白数据库和Excel文档")
        print("2. 添加操作员")
        print("3. 添加供应商")
        print("4. 添加仓库")
        print("5. 添加库存")
        print("6. 处理入库")
        print("7. 处理出库")
        print("8. 查看当前状态")
        print("9. 更新Excel报表")
        print("0. 退出系统")
        print("="*60)
    
    def interactive_menu(self):
        """交互式菜单"""
        while True:
            self.show_menu()
            choice = input("请选择操作 (0-9): ").strip()
            
            if choice == "0":
                print("👋 感谢使用仓库管理系统！")
                break
            elif choice == "1":
                self.create_blank_database()
                self.create_blank_excel()
                self.connect_database()
            elif choice == "2":
                name = input("请输入操作员姓名: ").strip()
                contact = input("请输入联系方式: ").strip()
                self.add_operator(name, contact)
            elif choice == "3":
                code = input("请输入供应商编号: ").strip()
                name = input("请输入供应商名称: ").strip()
                contact = input("请输入联系人: ").strip()
                phone = input("请输入联系方式: ").strip()
                self.add_supplier(code, name, contact, phone)
            elif choice == "4":
                name = input("请输入仓库名称: ").strip()
                operator = input("请输入操作员姓名: ").strip()
                manager = input("请输入负责人: ").strip()
                self.add_warehouse(name, operator, manager)
            elif choice == "5":
                code = input("请输入库存编号: ").strip()
                warehouse = input("请输入仓库名称: ").strip()
                quantity = int(input("请输入数量: ").strip())
                price = float(input("请输入单价: ").strip())
                self.add_inventory(code, warehouse, quantity, price)
            elif choice == "6":
                inbound_code = input("请输入入库编号: ").strip()
                inventory_code = input("请输入库存编号: ").strip()
                goods_code = input("请输入货物编号: ").strip()
                quantity = int(input("请输入数量: ").strip())
                name = input("请输入货物名称: ").strip()
                price = float(input("请输入单价: ").strip())
                supplier = input("请输入供应商: ").strip()
                self.process_inbound(inbound_code, inventory_code, goods_code, quantity, name, price, supplier)
            elif choice == "7":
                outbound_code = input("请输入出库编号: ").strip()
                inventory_code = input("请输入库存编号: ").strip()
                goods_code = input("请输入货物编号: ").strip()
                quantity = int(input("请输入数量: ").strip())
                name = input("请输入货物名称: ").strip()
                price = float(input("请输入单价: ").strip())
                self.process_outbound(outbound_code, inventory_code, goods_code, quantity, name, price)
            elif choice == "8":
                self.show_current_status()
            elif choice == "9":
                self.update_excel_report("手动更新")
            else:
                print("❌ 无效选择，请重新输入")
    
    def show_current_status(self):
        """显示当前状态"""
        print("\n" + "="*60)
        print("📊 当前系统状态")
        print("="*60)
        
        try:
            # 显示库存状态
            self.cursor.execute('''
                SELECT k.bianhao, k.cangkumingcheng, k.shuliang, k.danjia,
                       (k.shuliang * k.danjia) as 总价值
                FROM kucun k
                ORDER BY k.cangkumingcheng, k.bianhao
            ''')
            
            results = self.cursor.fetchall()
            if results:
                print("库存状态:")
                print(f"{'库存编号':<12} {'仓库名称':<12} {'数量':<8} {'单价':<10} {'总价值':<12}")
                print("-" * 60)
                for row in results:
                    print(f"{row[0]:<12} {row[1]:<12} {row[2]:<8} {row[3]:<10} {row[4]:<12}")
            else:
                print("暂无库存数据")
            
            # 显示仓库汇总
            self.cursor.execute('''
                SELECT c.cangkumingcheng, COUNT(k.bianhao) as 库存种类,
                       SUM(k.shuliang) as 总数量, SUM(k.shuliang * k.danjia) as 总价值
                FROM cangku c
                LEFT JOIN kucun k ON c.cangkumingcheng = k.cangkumingcheng
                GROUP BY c.cangkumingcheng
                ORDER BY c.cangkumingcheng
            ''')
            
            results = self.cursor.fetchall()
            if results:
                print("\n仓库汇总:")
                print(f"{'仓库名称':<12} {'库存种类':<8} {'总数量':<8} {'总价值':<12}")
                print("-" * 50)
                for row in results:
                    print(f"{row[0]:<12} {row[1]:<8} {row[2]:<8} {row[3]:<12}")
            else:
                print("\n暂无仓库数据")
                
        except Exception as e:
            print(f"❌ 获取状态失败: {e}")

def main():
    """主函数"""
    print("🚀 仓库管理系统 - 便捷操作工具")
    print("="*60)
    
    # 创建工具实例
    tool = WarehouseManagerTool()
    
    # 启动交互式菜单
    tool.interactive_menu()
    
    # 关闭数据库连接
    tool.close_database()

if __name__ == "__main__":
    main() 